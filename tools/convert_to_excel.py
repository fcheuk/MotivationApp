import json
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime

# 读取JSON文件
with open('../MotivationApp/Resources/quotes.json', 'r', encoding='utf-8') as f:
    quotes = json.load(f)

# 创建Excel工作簿
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "名言警句"

# 设置表头
headers = ['ID', '内容', '作者', '分类', '是否收藏', '创建日期']
ws.append(headers)

# 设置表头样式
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=12)

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')

# 添加数据
for quote in quotes:
    ws.append([
        quote['id'],
        quote['content'],
        quote['author'],
        quote['categoryId'],
        '是' if quote['isFavorite'] else '否',
        quote['createdDate']
    ])

# 调整列宽
ws.column_dimensions['A'].width = 40
ws.column_dimensions['B'].width = 60
ws.column_dimensions['C'].width = 20
ws.column_dimensions['D'].width = 15
ws.column_dimensions['E'].width = 12
ws.column_dimensions['F'].width = 20

# 设置文本对齐
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    row[1].alignment = Alignment(wrap_text=True, vertical='top')  # 内容列自动换行

# 保存Excel文件
output_file = f'quotes_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
wb.save(output_file)
print(f'转换完成！文件已保存为: {output_file}')